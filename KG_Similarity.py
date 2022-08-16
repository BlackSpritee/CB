"""
代码说明：
        由于文本打印模块不方便提供，只公布主要的几个功能函数
        该代码支持计算两个表示图谱的电子文档的相似度，输出整体相似度和细节相似度文档
            用户可自定义sim_mode相似度计算模式（默认为编辑距离）；
            用户必须输入两个产品标准的的关系表的路径；
"""

#两种相似度计算模式 编辑距离为"edit_distance"，用bert语义相似度的话改为"bert"
sim_mode="edit_distance"
# 填写两个产品标准的的关系表的路径
product1_path = './通舱套管/带复板的通舱套管-1/relations.xlsx'
product2_path = './通舱套管/带复板的通舱套管-2/relations.xlsx'

from openpyxl import Workbook, load_workbook #解析excel的包
if sim_mode=="bert": #在语义相似度计算模式下，需要安装torch和transformers
    from transformers import BertTokenizer, BertModel
    import torch
    PATH = "../../bert-base-chinese"  #本地离线的bert预训练包 下载地址为https://huggingface.co/bert-base-chinese/tree/main
    max_length = 128  #最大截止长度
    tokenizer = BertTokenizer.from_pretrained(PATH) #分词
    bert = BertModel.from_pretrained(PATH) #导入预训练模型

def editDistance(word1: str, word2: str) -> int:
    ''' 计算两个字符串的编辑距离度量的相似度
    :param word1: 字符串1
    :param word2: 字符串2
    :return: 编辑距离衡量的相似度
    '''
    n1 = len(word1)
    n2 = len(word2)
    dp = [[0] * (n2 + 1) for _ in range(n1 + 1)]

    # init
    for j in range(1, n2 + 1):
        dp[0][j] = j
    for i in range(1, n1 + 1):
        dp[i][0] = i

    for i in range(1, n1 + 1):
        for j in range(1, n2 + 1):
            if word1[i - 1] == word2[j - 1]:
                dp[i][j] = dp[i - 1][j - 1]
            else:
                dp[i][j] = min(dp[i][j - 1], dp[i - 1][j], dp[i - 1][j - 1]) + 1
    len_max = max(len(word1), len(word2))
    return (len_max - dp[-1][-1]) / len_max

def bert_sim(word1: str, word2: str) -> int:
    ''' 计算两个字符串的bert_cls语义向量衡量的相似度
    :param word1: 字符串1
    :param word2: 字符串2
    :return: bert_cls语义向量衡量的相似度
    '''

    def tokenize(word, max_length, tokenizer): #将字词转化为对应的id
        indexed_tokens = tokenizer.encode(word, add_special_tokens=True)
        avai_len = len(indexed_tokens)
        while len(indexed_tokens) < max_length:
            indexed_tokens.append(0)  # 0 is id for [PAD]
        indexed_tokens = indexed_tokens[: max_length]
        indexed_tokens = torch.tensor(indexed_tokens).long().unsqueeze(0)  # (1, L)
        # Attention mask
        att_mask = torch.zeros(indexed_tokens.size()).long()  # (1, L)
        att_mask[0, :avai_len] = 1
        return indexed_tokens, att_mask

    tokens1, mask1 = tokenize(word1, max_length, tokenizer)
    tokens2, mask2 = tokenize(word2, max_length, tokenizer)

    emb1 = bert(tokens1, attention_mask=mask1).last_hidden_state[:, 0] #bert编码
    emb2 = bert(tokens2, attention_mask=mask2).last_hidden_state[:, 0]
    sim = torch.nn.functional.cosine_similarity(emb1, emb2, dim=-1) #基于cls向量的余弦相似度计算
    return sim


def Kg2kg_similarity(kg1, kg2):
    '''
    计算两个图谱（字典类型{label:name}）的相似度
    :param kg1: 图谱1的字典
    :param kg2: 图谱2的字典
    :return: 图谱相似度的字典
    '''
    similarity_dict = {}
    for key1 in kg1:
        flag = 0
        for key2 in kg2:
            if key1 == key2: #两个图谱中实体类别相同时
                flag = 1
                similarity_dict[key1] = keyvalue_similarity(kg1[key1], kg2[key2]) #计算两个图谱相同实体类别下，具体实体的相似度
        if flag == 0:
            similarity_dict[key1] = 0

    return similarity_dict


def keyvalue_similarity(value1, value2):
    '''
    计算两个kg对应的key的value（label list）的相似度
    :param value1:list
    :param value2:list
    :return:相似度list，list中的每一项为一个元组（与v1最相似v2的相似度，对应v2在value2中的元素下标index
    '''
    sim_list = []
    for v1 in value1:
        sim = 0
        flag = 0
        for index, v2 in enumerate(value2):
            if sim_mode=='bert':
                sim_now=bert_sim(v1,v2)
            else:
                sim_now = editDistance(v1, v2)
            if sim_now > sim: #记住与v1最相似的v2值和最相似时的v2的index
                sim = sim_now
                flag = index
        sim_list.append((sim, flag))
    return sim_list



def entity_key(path):
    '''
    对输入路径对应的电子表格进行解析，根据列名来读取实体词、关系词、和三元组
    :param path: str
    :return: entity_key, relation, triple ：dict，list，list
    '''
    entity_key = {}  # label_name 头尾实体都在里面
    relation = []
    # wb = Workbook()
    wb = load_workbook(path)
    ws = wb.active
    row = 2
    col = 1
    while ws.cell(1, col).value != 'label1':
        col += 1
    while ws.cell(row, col).value != None:
        if ws.cell(row, col).value in entity_key and str(ws.cell(row, col + 1).value) not in entity_key[
            ws.cell(row, col).value]:
            entity_key[ws.cell(row, col).value].append(str(ws.cell(row, col + 1).value))
        else:
            entity_key[ws.cell(row, col).value] = [str(ws.cell(row, col + 1).value)]
        row += 1

    row = 2
    while ws.cell(1, col).value != 'rel':
        col += 1
    while ws.cell(row, col).value != None:
        if ws.cell(row, col).value not in relation:
            relation.append(ws.cell(row, col).value)
        row += 1

    row = 2
    while ws.cell(1, col).value != 'label2':
        col += 1
    while ws.cell(row, col).value != None:
        if ws.cell(row, col).value in entity_key and str(ws.cell(row, col + 1).value) not in entity_key[
            ws.cell(row, col).value]:
            entity_key[ws.cell(row, col).value].append(str(ws.cell(row, col + 1).value))
        else:
            entity_key[ws.cell(row, col).value] = [str(ws.cell(row, col + 1).value)]
        row += 1

    col = 0
    while (True):
        col += 1
        if ws.cell(1, col).value == 'name1':
            sub_name_col = col
        if ws.cell(1, col).value == 'rel':
            rel_col = col
        if ws.cell(1, col).value == 'name2':
            obj_name_col = col
            break
    triple = []
    row = 1
    while ws.cell(row, sub_name_col).value != None:
        row += 1
        if ws.cell(row, obj_name_col).value != None and ws.cell(row, rel_col).value != None:
            triple.append(
                [ws.cell(row, sub_name_col).value, ws.cell(row, rel_col).value, ws.cell(row, obj_name_col).value])
    return entity_key, relation, triple

